from src import satnogs_api, satnogs_selection
import openpyxl  # excel
from datetime import datetime
import itertools
from os import path

TABSNAME = {0: "allSatellite", 1: "filteredSatellite", 2: "sortedSatellite", 3: "TLE"}
EXCEL_DIR = 'D:\\satnogs_satellites_' + str(datetime.now().date()) + '.xlsx'
TLE_DIR = 'D:\\tle.save'


def newWorkbook() -> openpyxl.Workbook:
    """
    :return: A new Excel file object
    """
    return openpyxl.Workbook()


def newTab(wb: openpyxl.Workbook, tabName: str, tabIndex: int = 0) -> openpyxl.Workbook.worksheets:
    """
    :param wb: Excel file object
    :param tabName: Name of the current Excel tab
    :param tabIndex: Index of the current Excel tab, start from 0
    :return:
    """
    return wb.create_sheet(tabName, tabIndex)


def addHeader(ws: openpyxl.Workbook.worksheets) -> None:
    """
    Add these fix header into current Excel tab
    :param ws: A specific worksheet object within an Excel file object
    :return:
    """
    ws["A1"] = "Name"
    ws["B1"] = "Description"
    ws["C1"] = "Norad_cat_id"
    ws["D1"] = "Service"
    ws["E1"] = "Mode"
    ws["F1"] = "Baud Rate"
    ws["G1"] = "Timestamp"


def exportData(ws: openpyxl.Workbook.worksheets, result: [dict]) -> None:
    """
    Export Satellite information on to current Excel worksheet
    :param ws: A specific worksheet object within an Excel file object
    :param result: Filtered Satellite data from API
    :return:
    """

    for r in range(2, len(result) + 2):
        entry = result[r - 2]

        ws.cell(r, 1, entry["name"])
        ws.cell(r, 2, entry["description"])
        ws.cell(r, 3, entry["norad_cat_id"])
        ws.cell(r, 4, entry["service"])
        ws.cell(r, 5, entry["mode"])
        ws.cell(r, 6, entry["baud"])
        ws.cell(r, 7, entry["time"])


def saveTLE(TLE: [dict]) -> [dict]:
    # TODO: complete this function
    """
    this function should save dict into a raw text file in a format that is easy to read from
    the saved text file will be later loaded and convert back into the original dict format
    :param TLE: list of raw TLE data from API
    :return: return param
    """

    f = open(TLE_DIR, 'w')
    curr_time = datetime.now()
    f.write(str(curr_time) + "\n")
    for line in TLE:
        for v in line.values():
            f.write(str(v) + "\n")
    f.close()

    return TLE


def loadTLE(fileDir: str = TLE_DIR) -> [dict]:
    # TODO: complete this function
    """
    this function should load TLE from text file and convert back to original dict formatting
    :param fileDir: directory for the text file
    :return:
    """
    TLE = []

    try:
        f = open(fileDir, 'r')
    except FileNotFoundError:
        print("fileNotFound")
        rawData = satnogs_selection.tleFilter(
            satnogs_selection.sortMostRecent(satnogs_selection.satelliteFilter(satnogs_api.getSatellites())))
        return saveTLE(rawData)
    else:
        lines = f.readlines()
        saved_time = lines[0].strip()
        date_time_obj = datetime.strptime(saved_time, '%Y-%m-%d %H:%M:%S.%f')
        curr_time = datetime.now()

        if (curr_time - date_time_obj).days >= 1:
            print("fileOutDated")
            rawData = satnogs_selection.tleFilter(
                satnogs_selection.sortMostRecent(satnogs_selection.satelliteFilter(satnogs_api.getSatellites())))
            return saveTLE(rawData)
        else:
            print("readingExistingFile")
            repeatPattern = 6
            for line in range(1, len(lines), repeatPattern):
                keys = ['tle0', 'tle1', 'tle2', 'tle_source', 'norad_cat_id', 'updated']
                tle = dict()
                for k in range(repeatPattern):
                    if k == 4:
                        tle[keys[k]] = int(lines[line + k].strip())
                    else:
                        tle[keys[k]] = lines[line+k].strip()
                TLE.append(tle)

    return TLE


def saveFile(wb: openpyxl.Workbook, dir: str) -> None:
    """
    :param wb: Excel file object
    :param dir: Directory to save Excel file
    :return:
    """
    wb.save(dir)


def exportTab(wb: openpyxl.Workbook, tab: int, result: [dict]) -> None:
    #  export data from dict to excel tab
    if not result:  # if list is empty, do nothing
        return

    ws = newTab(wb, TABSNAME[tab], tab)
    addHeader(ws)
    exportData(ws, result)


def exportTLE(wb: openpyxl.Workbook, tab: int, result: [dict]) -> None:
    """
    :param wb:
    :param tab:
    :param result:
    :return:
    """

    if not result:  # if list is empty, do nothing
        return

    ws = newTab(wb, "TLE", tab)

    ws["A1"] = "tle0"
    ws["B1"] = "tle1"
    ws["C1"] = "tle2"
    ws["D1"] = "tle_source"
    ws["E1"] = "norad_cat_id"
    ws["F1"] = "updated"

    for r in range(0, len(result)):
        # print(result[r])
        for c in range(0, len(result[r])):
            ws.cell(r + 2, c + 1, list(result[r].values())[c])


# driver

# allSatellite = satnogs_api.getSatellites()
# filteredSatellite = satnogs_selection.satelliteFilter(allSatellite)
# sortedSatellite = satnogs_selection.sortMostRecent(filteredSatellite)
# TLE = satnogs_selection.tleFilter(sortedSatellite)
#
# wb = newWorkbook()
# exportTab(wb, 0, allSatellite)
# exportTab(wb, 1, filteredSatellite)
# exportTab(wb, 2, sortedSatellite)
# exportTLE(wb, 3, TLE)
# saveFile(wb, EXCEL_DIR)
# wb.close()


# # saveTLE(TLE)
#
# loaded = loadTLE('TLE_data.txt')
#
# # for i in range(0, len(TLE)):
# #     assert TLE[i] == loaded[i]